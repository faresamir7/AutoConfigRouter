"""This module provides the AutoConfigRouter CLI."""
# AutoConfigRouter/cli.py

from typing import Optional
from pathlib import Path
from openpyxl import load_workbook

import openpyxl
import typer
import AutoConfigRouter.autoconfigrouter


from AutoConfigRouter import __app_name__, __version__, ERRORS, config

app = typer.Typer()

@app.command()
def getRulesCisco(
    ip_addr: str = typer.Option(
        "192.168.1.1",
        "--ip-address",
        "-ip",
        prompt="Set ip address:",
    ),
    login: str = typer.Option(
        "admin",
        "--login",
        "-l",
        prompt="Set login credential:",
    ),
    password: str = typer.Option(
        "password",
        "--passwd",
        "-p",
        prompt="Set password:",
    ),
) -> None:
    """Get the rules from a Cisco FMC."""
    AutoConfigRouter.autoconfigrouter.getRulesCisco(login,password,ip_addr)

@app.command()
def getDevicesCisco(
    ip_addr: str = typer.Option(
        "192.168.1.1",
        "--ip-address",
        "-ip",
        prompt="Set ip address:",
    ),
    login: str = typer.Option(
        "admin",
        "--login",
        "-l",
        prompt="Set login credential:",
    ),
    password: str = typer.Option(
        "password",
        "--passwd",
        "-p",
        prompt="Set password:",
    ),
) -> None:
    """Get the devices from a Cisco FMC."""
    AutoConfigRouter.autoconfigrouter.getDevicesCisco(login,password,ip_addr)

@app.command()
def addFTDCisco(
    ip_addr: str = typer.Option(
        "192.168.1.1",
        "--ip-address",
        "-ip",
        prompt="Set ip address:",
    ),
    login: str = typer.Option(
        "admin",
        "--login",
        "-l",
        prompt="Set login credential:",
    ),
    password: str = typer.Option(
        "password",
        "--passwd",
        "-p",
        prompt="Set password:",
    ),
    file: str = typer.Option(
        r"C:\Users\User\Documents\spreadsheet.xlsx",
        "--file",
        "-f",
        prompt="Set file location:",
    ),
) -> None:
    """Add the FTD device to the target FMC."""
    wb = openpyxl.Workbook()
    wb = load_workbook(filename = file, read_only=1)
    ws = wb.active
    ftdSet = []
    ftd = []
    for i in range(2, ws.max_row+1):
        ftd.clear()
        for j in range(1, ws.max_column+1):
            data = ws.cell(row=i, column=j)
            ftd.append(data.value)
        ftdSet.append(ftd.copy())
    AutoConfigRouter.autoconfigrouter.addFTDCisco(ftdSet,login,password,ip_addr)

@app.command()
def getUsersCisco(
    ip_addr: str = typer.Option(
        "192.168.1.1",
        "--ip-address",
        "-ip",
        prompt="Set ip address:",
    ),
    login: str = typer.Option(
        "cisco",
        "--login",
        "-l",
        prompt="Set login credential:",
    ),
    password: str = typer.Option(
        "cisco123",
        "--passwd",
        "-p",
        prompt="Set password:",
    ),
) -> None:
    """Get the users from a Cisco FMC."""
    AutoConfigRouter.autoconfigrouter.getUsersCisco(login,password,ip_addr)

@app.command()
def uploadRulesCisco(
    ip_address: str = typer.Option(
        "192.168.1.1",
        "--ip-address",
        "-ip",
        prompt="Set ip address:",
    ),
    login: str = typer.Option(
        "cisco",
        "--login",
        "-l",
        prompt="Set login credential:",
    ),
    password: str = typer.Option(
        "cisco123",
        "--passwd",
        "-p",
        prompt="Set password:",
    ),
    containerName: str = typer.Option(
        "Test Policy",
        "--container",
        "-c",
        prompt="Set Policy Name:"
    ),
    file: str = typer.Option(
        r"C:\Users\User\Documents\spreadsheet.xlsx",
        "--file",
        "-f",
        prompt="Set file location:",
    ),
) -> None:
    """Upload rulesets from an excel file to a Cisco FMC."""
    if(AutoConfigRouter.autoconfigrouter.getPolicyCisco(login,password,ip_address,containerName) == False):
        return
    wb = openpyxl.Workbook()
    wb = load_workbook(filename = file, read_only=1)
    ws = wb.active
    ruleset = []
    rule = []
    for i in range(2, ws.max_row+1):
        rule.clear()
        for j in range(1, ws.max_column+1):
            data = ws.cell(row=i, column=j)
            rule.append(data.value)
        ruleset.append(rule.copy())
    AutoConfigRouter.autoconfigrouter.addRulesCisco(ruleset,login,password,ip_address,containerName)

@app.command()
def getPolicyCisco(
    ip_address: str = typer.Option(
        "192.168.1.1",
        "--ip-address",
        "-ip",
        prompt="Set ip address:",
    ),
    login: str = typer.Option(
        "cisco",
        "--login",
        "-l",
        prompt="Set login credential:",
    ),
    password: str = typer.Option(
        "cisco123",
        "--passwd",
        "-p",
        prompt="Set password:",
    ),
    containerName: str = typer.Option(
        "Test Policy",
        "--container",
        "-c",
        prompt="Set Policy Name:"
    ),
) -> None:
    """Get a container's information from a Cisco FMC."""
    AutoConfigRouter.autoconfigrouter.getPolicyCisco(login,password,ip_address,containerName)

@app.command()
def uploadForti(
    ip_address: str = typer.Option(
        "192.168.1.1",
        "--ip-address",
        "-ip",
        prompt="Set ip address:",
    ),
    api_key: str = typer.Option(
        "gHmt8z6rx49Qy58Q9mkNqn377jxmH8",
        "--key",
        "-k",
        prompt="Set API-User Key:",
    ),
    file: str = typer.Option(
        r"C:\Users\User\Documents\spreadsheet.xlsx",
        "--file",
        "-f",
        prompt="Set file location:",
    ),
) -> None:
    """Upload a rulebook to a Fortinet firewall."""
    wb = openpyxl.Workbook()
    wb = load_workbook(filename = file, read_only=1)
    ws = wb.active
    ruleset = []
    rule = []
    for i in range(2, ws.max_row+1):
        rule.clear()
        for j in range(1, ws.max_column+1):
            data = ws.cell(row=i, column=j)
            if(data.value != 'None'):
                rule.append(data.value)
            elif(data.value != 'None' and j<=ws.max_column+1):
                rule.append(None)
        ruleset.append(rule.copy())
    #TODO make a create function for Fortinet in spare time

@app.command()
def getForti(
    ip_addr: str = typer.Option(
        "192.168.1.1",
        "--ip-address",
        "-ip",
        prompt="Set ip address:",
    ),
    login: str = typer.Option(
        "cisco",
        "--login",
        "-l",
        prompt="Set login credential:",
    ),
    password: str = typer.Option(
        "cisco123",
        "--passwd",
        "-p",
        prompt="Set password:",
    ),
) -> None:
    """Get the rules from a fortinet firewall."""
    AutoConfigRouter.autoconfigrouter.getRulesFortigate(login,password,ip_addr)

@app.command()
def validate(
    document_path: str = typer.Option(
        r"C:\Users\User\Documents\spreadsheet.xlsx",
        "--document",
        "-d",
        prompt="Select valid document to read ruleset from:",
    ),
) -> None:
    """Read rules from an excel file."""
    wb = openpyxl.Workbook()
    wb = load_workbook(filename = document_path, read_only=1)
    ws = wb.active
    for i in range(2, ws.max_row+1):
        for j in range(1, ws.max_column+1):
            cell_obj = ws.cell(row=i, column=j)
            cell_type = ws.cell(row=1, column=j)
            print(cell_type.value,": ", cell_obj.value, end="\n")
        print("\n")

def _version_callback(value: bool) -> None:
    if value:
        typer.echo(f"{__app_name__} v{__version__}")
        raise typer.Exit()

@app.callback()
def main(
    version: Optional[bool] = typer.Option(
        None,
        "--version",
        "-v",
        help="Show the application's version and exit.",
        callback=_version_callback,
        is_eager=True,
    )
) -> None:
    return